"""扫描报告导出单元测试。

覆盖 ``fuscan.export.report`` 模块的 PDF/Excel 二进制导出与 ``save_report``
按扩展名分发的文件保存逻辑。测试用例从 ``test_scanner.py`` 迁移而来，
原 ``ScanReport.to_pdf``/``to_excel``/``save_report`` 方法已拆分到本模块。
"""

from __future__ import annotations

from pathlib import Path

import pytest

from fuscan.export.report import export_excel, export_pdf, save_report
from fuscan.rules.model import Severity
from fuscan.scanner import ScanReport, ScanResult
from fuscan.scanner.result import RuleHit, ScanStats


def _build_report(tmp_path: Path) -> ScanReport:
    """构造测试报告：3 个文件命中 2 条规则，分属 WARNING/CRITICAL 两个等级。"""
    results = (
        ScanResult(
            path=tmp_path / "secret.txt" / "a.txt",
            size=10,
            hits=(
                RuleHit("敏感文件名", Severity.WARNING, "d1", match_count=1),
                RuleHit("密钥内容", Severity.CRITICAL, "d2", match_count=2),
            ),
        ),
        ScanResult(
            path=tmp_path / "secret.txt" / "b.txt",
            size=20,
            hits=(RuleHit("密钥内容", Severity.CRITICAL, "d3", match_count=3),),
        ),
        ScanResult(path=tmp_path / "clean.txt", size=0, hits=()),
    )
    stats = ScanStats(
        total_files=3,
        scanned_files=3,
        matched_files=2,
        skipped_files=0,
        errors=0,
        duration_seconds=0.5,
        total_matches=6,
    )
    return ScanReport(root=tmp_path, results=results, stats=stats)


class TestExportPdf:
    def test_export_pdf_returns_bytes_with_header(self, tmp_path: Path) -> None:
        """export_pdf 应返回 PDF 二进制数据，以 %PDF- 开头。"""
        report = _build_report(tmp_path)
        data = export_pdf(report)
        assert isinstance(data, bytes)
        assert data[:5] == b"%PDF-"

    def test_export_pdf_empty_hits(self, tmp_path: Path) -> None:
        """空命中报告也能生成 PDF，仍以 %PDF- 开头。"""
        report = ScanReport(root=tmp_path, results=(), stats=ScanStats())
        data = export_pdf(report)
        assert data[:5] == b"%PDF-"

    def test_export_pdf_contains_keywords(self, tmp_path: Path) -> None:
        """PDF 文本流应包含 PDF 结构标记（CID 字体下中文不可直接 grep）。"""
        report = _build_report(tmp_path)
        data = export_pdf(report)
        # PDF 中文字以 CID 编码无法直接 grep，但 PDF 结构标记应可见
        assert b"/Type /Catalog" in data or b"/Pages" in data

    def test_export_pdf_truncates_long_detail(self, tmp_path: Path) -> None:
        """iter-138：超长 detail 应被截断到 200 字符 + 省略号，避免 LayoutError。

        原始 bug：未截断时单 cell 换行后行高可达 4972pt > A4 页面可用高度，
        触发 ``LayoutError: Table N rows x M cols too large on page``。
        """
        # 构造 5000 字符的超长 detail（模拟 base64 编码内容）
        long_detail = "A" * 5000
        results = (
            ScanResult(
                path=tmp_path / "leak.txt",
                size=100,
                hits=(
                    RuleHit(
                        rule_name="密钥泄漏",
                        severity=Severity.CRITICAL,
                        detail=long_detail,
                        match_count=1,
                    ),
                ),
            ),
        )
        stats = ScanStats(
            total_files=1,
            scanned_files=1,
            matched_files=1,
            total_matches=1,
        )
        report = ScanReport(root=tmp_path, results=results, stats=stats)
        # 不应抛 LayoutError，应正常生成 PDF
        data = export_pdf(report)
        assert data[:5] == b"%PDF-"

    def test_export_pdf_with_engine_set(self, tmp_path: Path) -> None:
        """result.engine 非空时 PDF 表格应含「引擎」列且正常生成（OCR 回退场景）。"""
        results = (
            ScanResult(
                path=tmp_path / "scan.pdf",
                size=100,
                hits=(RuleHit("密钥", Severity.CRITICAL, "d", match_count=1),),
                engine="pypdfium2 + rapidocr-json",
            ),
        )
        report = ScanReport(
            root=tmp_path,
            results=results,
            stats=ScanStats(total_files=1, scanned_files=1, matched_files=1, total_matches=1),
        )
        # 不应抛异常，应正常生成 PDF
        data = export_pdf(report)
        assert data[:5] == b"%PDF-"

    def test_truncate_text_helper(self) -> None:
        """iter-138：_truncate_text 辅助函数应正确截断超长文本。"""
        from fuscan.export.report import _truncate_text

        # 短文本原样返回
        assert _truncate_text("short") == "short"
        # 边界：恰好等于阈值
        assert _truncate_text("A" * 200) == "A" * 200
        # 超长文本截断 + 省略号
        result = _truncate_text("A" * 500)
        assert len(result) == 203  # 200 + "..."
        assert result.endswith("...")
        # 自定义阈值
        result_custom = _truncate_text("A" * 100, max_chars=50)
        assert len(result_custom) == 53
        assert result_custom.endswith("...")


class TestExportExcel:
    def test_export_excel_returns_zip_archive(self, tmp_path: Path) -> None:
        """export_excel 应返回 xlsx 二进制数据（zip 格式，PK 开头）。"""
        report = _build_report(tmp_path)
        data = export_excel(report)
        assert isinstance(data, bytes)
        # xlsx 是 zip 压缩包，开头为 PK\x03\x04
        assert data[:2] == b"PK"

    def test_export_excel_empty_hits(self, tmp_path: Path) -> None:
        """空命中报告也能生成 Excel。"""
        report = ScanReport(root=tmp_path, results=(), stats=ScanStats())
        data = export_excel(report)
        assert data[:2] == b"PK"

    def test_export_excel_roundtrip(self, tmp_path: Path) -> None:
        """生成的 xlsx 应能被 openpyxl 读回，且工作表名称正确。"""
        from openpyxl import load_workbook

        report = _build_report(tmp_path)
        data = export_excel(report)
        import io as _io

        wb = load_workbook(_io.BytesIO(data))
        assert "扫描汇总" in wb.sheetnames
        assert "命中明细" in wb.sheetnames
        # 命中明细表头应在第 1 行，含「解析引擎」列
        headers = [c.value for c in wb["命中明细"][1]]
        assert headers == ["路径", "大小", "严重等级", "规则", "描述", "匹配数", "解析引擎", "详情"]

    def test_export_excel_engine_from_result(self, tmp_path: Path) -> None:
        """result.engine 非空时应写入运行期引擎名（如 OCR 回退）。"""
        import io as _io

        from openpyxl import load_workbook

        results = (
            ScanResult(
                path=tmp_path / "scan.pdf",
                size=100,
                hits=(RuleHit("密钥", Severity.CRITICAL, "d", match_count=1),),
                engine="pypdfium2 + rapidocr-json",
            ),
        )
        report = ScanReport(
            root=tmp_path,
            results=results,
            stats=ScanStats(total_files=1, scanned_files=1, matched_files=1, total_matches=1),
        )
        data = export_excel(report)
        wb = load_workbook(_io.BytesIO(data))
        # 第 2 行第 7 列（解析引擎）应为运行期引擎名
        assert wb["命中明细"].cell(row=2, column=7).value == "pypdfium2 + rapidocr-json"

    def test_export_excel_engine_fallback(self, tmp_path: Path) -> None:
        """result.engine 为空时回退到 engine_for_extension 静态映射。"""
        import io as _io

        from openpyxl import load_workbook

        from fuscan.scanner._helpers import engine_for_extension

        report = _build_report(tmp_path)
        data = export_excel(report)
        wb = load_workbook(_io.BytesIO(data))
        # _build_report 全部 .txt 文件，engine 为空 → 回退 engine_for_extension("txt")
        expected = engine_for_extension("txt")
        assert wb["命中明细"].cell(row=2, column=7).value == expected


class TestSaveReport:
    def test_save_report_csv(self, tmp_path: Path) -> None:
        """save_report 按 .csv 扩展名写入 UTF-8 文本。"""
        report = _build_report(tmp_path)
        target = tmp_path / "out.csv"
        save_report(report, target)
        content = target.read_text(encoding="utf-8")
        # iter-89：CSV 列新增 archive_path/inner_path 标识压缩包内部条目
        assert content.startswith("path,archive_path,inner_path,size,severity,rule,description,match_count,detail")

    def test_save_report_json(self, tmp_path: Path) -> None:
        """save_report 按 .json 扩展名写入 JSON 文本。"""
        import json as _json

        report = _build_report(tmp_path)
        target = tmp_path / "out.json"
        save_report(report, target)
        data = _json.loads(target.read_text(encoding="utf-8"))
        assert data["root"] == str(tmp_path)

    def test_save_report_txt_fallback(self, tmp_path: Path) -> None:
        """save_report 对 .txt 扩展名按 text 格式写入。"""
        report = _build_report(tmp_path)
        target = tmp_path / "out.txt"
        save_report(report, target)
        content = target.read_text(encoding="utf-8")
        assert "扫描路径:" in content

    def test_save_report_pdf(self, tmp_path: Path) -> None:
        """save_report 按 .pdf 扩展名写入二进制。"""
        report = _build_report(tmp_path)
        target = tmp_path / "out.pdf"
        save_report(report, target)
        data = target.read_bytes()
        assert data[:5] == b"%PDF-"

    def test_save_report_xlsx(self, tmp_path: Path) -> None:
        """save_report 按 .xlsx 扩展名写入二进制。"""
        report = _build_report(tmp_path)
        target = tmp_path / "out.xlsx"
        save_report(report, target)
        data = target.read_bytes()
        assert data[:2] == b"PK"

    def test_save_report_unknown_ext_falls_back_to_text(self, tmp_path: Path) -> None:
        """save_report 对未知扩展名（非 csv/json/pdf/xlsx）按 text 写入。"""
        report = _build_report(tmp_path)
        target = tmp_path / "out.log"
        save_report(report, target)
        content = target.read_text(encoding="utf-8")
        assert "扫描路径:" in content


class TestScanReportJsonRoundtrip:
    """iter-123：ScanReport.to_json/from_json 反序列化回环测试。"""

    def test_from_json_roundtrip_preserves_basic_fields(self, tmp_path: Path) -> None:
        """to_json → from_json 应保留 root/cancelled/stats 基本字段。"""
        report = _build_report(tmp_path)
        restored = ScanReport.from_json(report.to_json())
        assert restored.root == report.root
        assert restored.cancelled == report.cancelled
        assert restored.stats.total_files == report.stats.total_files
        assert restored.stats.scanned_files == report.stats.scanned_files
        assert restored.stats.matched_files == report.stats.matched_files
        assert restored.stats.total_matches == report.stats.total_matches

    def test_from_json_roundtrip_preserves_hits(self, tmp_path: Path) -> None:
        """to_json → from_json 应保留命中结果（路径/大小/规则命中）。"""
        report = _build_report(tmp_path)
        restored = ScanReport.from_json(report.to_json())
        assert len(restored.hits) == len(report.hits)
        # 比较第一个命中文件的路径与规则数
        orig_hit = report.hits[0]
        restored_hit = restored.hits[0]
        assert restored_hit.path == orig_hit.path
        assert restored_hit.size == orig_hit.size
        assert len(restored_hit.hits) == len(orig_hit.hits)

    def test_from_json_roundtrip_preserves_severity(self, tmp_path: Path) -> None:
        """to_json → from_json 应保留规则严重等级（Severity 枚举）。"""
        report = _build_report(tmp_path)
        restored = ScanReport.from_json(report.to_json())
        # _build_report 含 WARNING 和 CRITICAL 两个等级
        severities = {h.severity for r in restored.hits for h in r.hits}
        assert Severity.WARNING in severities
        assert Severity.CRITICAL in severities

    def test_from_json_roundtrip_preserves_match_count(self, tmp_path: Path) -> None:
        """to_json → from_json 应保留 match_count（区分命中规则数与匹配条数）。"""
        report = _build_report(tmp_path)
        restored = ScanReport.from_json(report.to_json())
        assert len(report.hits) == len(restored.hits)
        for orig, restored_hit in zip(report.hits, restored.hits, strict=True):
            assert len(orig.hits) == len(restored_hit.hits)
            for orig_rule, restored_rule in zip(orig.hits, restored_hit.hits, strict=True):
                assert restored_rule.match_count == orig_rule.match_count

    def test_from_json_roundtrip_preserves_match_texts_tuple(self, tmp_path: Path) -> None:
        """to_json → from_json 应将 list 转回 tuple（asdict 将 tuple 序列化为 list）。"""
        from fuscan.scanner.result import RuleHit

        # 构造带 match_texts 的 RuleHit
        report = ScanReport(
            root=tmp_path,
            results=(
                ScanResult(
                    path=tmp_path / "a.txt",
                    size=10,
                    hits=(
                        RuleHit(
                            rule_name="test",
                            severity=Severity.WARNING,
                            detail="d",
                            match_texts=("pwd", "password"),
                        ),
                    ),
                ),
            ),
            stats=ScanStats(total_files=1, scanned_files=1, matched_files=1),
        )
        restored = ScanReport.from_json(report.to_json())
        assert restored.hits[0].hits[0].match_texts == ("pwd", "password")
        assert isinstance(restored.hits[0].hits[0].match_texts, tuple)

    def test_from_json_invalid_json_raises(self) -> None:
        """非法 JSON 应抛 ValueError（json.JSONDecodeError 的父类）。"""
        with pytest.raises(ValueError):
            ScanReport.from_json("not valid json")

    def test_from_json_non_dict_raises(self) -> None:
        """JSON 顶层非字典应抛 ValueError。"""
        with pytest.raises(ValueError, match="顶层必须是字典"):
            ScanReport.from_json("[1, 2, 3]")

    def test_from_json_empty_hits(self, tmp_path: Path) -> None:
        """空命中报告也能序列化/反序列化。"""
        report = ScanReport(root=tmp_path, results=(), stats=ScanStats())
        restored = ScanReport.from_json(report.to_json())
        assert restored.hits == ()
        assert restored.stats.total_files == 0

    def test_from_json_invalid_severity_falls_back_to_info(self, tmp_path: Path) -> None:
        """未知 severity 字符串回退为 INFO（不抛异常）。"""
        import json as _json

        data: dict[str, object] = {
            "root": str(tmp_path),
            "stats": {"total_files": 0},
            "cancelled": False,
            "hits": [
                {
                    "path": str(tmp_path / "a.txt"),
                    "size": 10,
                    "rules": [
                        {
                            "rule_name": "test",
                            "severity": "unknown_level",
                            "detail": "d",
                        }
                    ],
                }
            ],
        }
        restored = ScanReport.from_json(_json.dumps(data))
        assert restored.hits[0].hits[0].severity == Severity.INFO

    def test_from_json_perf_summary_not_persisted(self, tmp_path: Path) -> None:
        """perf_summary 不持久化（运行时统计重启后无意义）。"""
        report = _build_report(tmp_path)
        restored = ScanReport.from_json(report.to_json())
        assert restored.stats.perf_summary is None
